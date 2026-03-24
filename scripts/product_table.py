# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

DARK_BLUE   = "1E3A5F"
MID_BLUE    = "2C5F8A"
LIGHT_BLUE  = "D6E4F0"
ACCENT_BLUE = "4A90D9"
ACCENT_GREEN= "27AE60"
ACCENT_ORG  = "E67E22"
ACCENT_PURP = "8E44AD"
LIGHT_GREEN = "D5F5E3"
LIGHT_ORG   = "FDEBD0"
LIGHT_PURP  = "E8DAEF"
LIGHT_GRAY  = "F2F3F4"
WHITE       = "FFFFFF"
DARK_GRAY   = "2C3E50"

def fill(c): return PatternFill("solid", fgColor=c)
def fnt(bold=False, size=10, color="000000", italic=False):
    return Font(name="微软雅黑", bold=bold, size=size, color=color, italic=italic)
def cal(h=True): return Alignment(horizontal="center" if h else "left", vertical="center", wrap_text=True)
def lal(): return Alignment(horizontal="left", vertical="center", wrap_text=True)
def bdr(style="thin"):
    s = Side(style=style, color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def merge_write(ws, row, col_start, col_end, value, bg, font_obj, aln_obj, height=30):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=value)
    c.fill = fill(bg); c.font = font_obj; c.alignment = aln_obj; c.border = bdr()
    ws.row_dimensions[row].height = height

def write_sec_header(ws, row, text, bg, cols=5):
    merge_write(ws, row, 1, cols, "  " + text, bg, fnt(True, 13, WHITE), cal(), 28)
    return row + 1

def write_prod_row(ws, row, name, desc, example, note, bg):
    for col, val in [(1,""), (2,name), (3,desc), (4,example), (5,note)]:
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg); c.border = bdr()
        c.font = fnt(True,10,DARK_GRAY) if col==2 else fnt(False,10,DARK_GRAY)
        c.alignment = lal()
    ws.row_dimensions[row].height = 48

# ===================== Sheet 1 =====================
ws1 = wb.active
ws1.title = "产品总览"
for i, w in enumerate([4,22,40,20,30], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

merge_write(ws1, 1, 1, 5,
    "上海基数运维 · 合作伙伴产品手册",
    DARK_BLUE, fnt(True,18,WHITE), cal(), 50)
merge_write(ws1, 2, 1, 5,
    "机房基础设施 · 服务器与算力 · 企业系统 · 网络安全 · 一站式解决方案",
    MID_BLUE, fnt(False,10,WHITE,italic=True), cal(), 22)

r = 4

# -- 1. 机房基础设施 --
r = write_sec_header(ws1, r, "一、机房基础设施", DARK_BLUE)
for item in [
    ("机柜（服务器机柜）",
     "用于存放服务器、网络设备、UPS等核心设备的标准化金属柜体。常见规格为42U，宽度19英寸。",
     "例：某药企规划中期需50台服务器，首期先采购20台，配2个机柜加预留扩展位",
     "选型注意：散热方式（前进风/后出风）、承重（大于等于1000kg）、PDU插排类型",
     LIGHT_BLUE),
    ("PDU（电源分配单元）",
     "机柜内部专业级电源插排，相比普通插线板可实时监测电流功率，具备过载保护与远程控制功能。",
     "例：服务器双电源接入PDU，两路市电加UPS，实现供电冗余",
     "可精确统计每台设备功耗，远程一键断电重置故障设备",
     LIGHT_BLUE),
    ("精密空调（机房空调）",
     "为机房提供恒温恒湿（通常22正负2度，湿度45-65%）的制冷设备，普通家用空调无法满足7x24小时运行需求。",
     "例：某生物药企机房60平方米，配置2台40kW精密空调，N加1冗余",
     "制冷量选型约为总发热量乘以1.2（冗余系数），需评估气流组织",
     LIGHT_BLUE),
    ("UPS不间断电源",
     "市电中断时立即接管供电，保障服务器正常关机或切换到发电机，避免数据丢失和硬件损坏。",
     "例：药企ERP服务器配置30kVA UPS，可支撑30分钟供所有设备安全关机",
     "功率选型等于所有设备总功率乘以1.3（峰值加冗余），电池续航按实际需求配置",
     LIGHT_BLUE),
    ("环境监控（动环监控）",
     "对机房温湿度、烟感、漏水、门禁、供电等集中监控，超阈值自动告警（短信/电话/微信）。",
     "例：某客户凌晨UPS故障，监控系统3秒内拨出报警电话，运维人员15分钟到场",
     "建设期提前布线，后期加装成本翻3倍",
     LIGHT_BLUE),
]:
    write_prod_row(ws1, r, *item); r += 1
r += 2

# -- 2. 服务器与算力 --
r = write_sec_header(ws1, r, "二、服务器与算力设备", MID_BLUE)
for item in [
    ("标准服务器（1U/2U/4U机架式）",
     "数据中心主流服务器形态，安装在机柜中。1U最薄（追求密度），2U空间适中（扩展性好），4U最高（GPU服务器/存储型）。",
     "例：生物药企LIMS系统，预计5年内承载100用户，选2台2U服务器做HA双机热备",
     "品牌：戴尔/HPE/浪潮/新华三；BIOS设置、BMC固件、RAID配置由我们负责",
     LIGHT_ORG),
    ("AI服务器（GPU算力服务器）",
     "搭载NVIDIA A100/H100/H20等加速卡，支撑AI推理、模型训练、科学计算。药企可用于药物分子模拟、晶型预测、文献AI分析。",
     "例：某biotech部署AI辅助化合物筛选，一次性筛选10万个分子，效率提升200倍",
     "注意显存大小（40G/80G/640G）、NVLink互联带宽、多卡扩展性",
     LIGHT_ORG),
    ("存储服务器（NAS/SAN/DAS）",
     "专门用于海量数据存储。NAS（文件级共享，适合20台以内）；SAN（块存储，高性能数据库）；DAS（直连，性价比首选）。",
     "例：药企GMP要求数据保存大于等于5年，历史实验数据约200TB，配置2台存储服务器做RAID备份",
     "选型核心看：容量、IOPS、吞吐量、冗余架构",
     LIGHT_ORG),
    ("一体机（超融合/数据库一体机）",
     "将计算、存储、网络、虚拟化整合到一台设备，3小时即可交付使用，大幅降低运维复杂度。",
     "例：某药企初创公司IT团队仅1人，采购超融合一体机承载ERP加CRM加文件共享，首周即上线",
     "适合预算有限但追求快速部署的中小药企，首选国产品牌（华为/新华三/深信服）",
     LIGHT_ORG),
]:
    write_prod_row(ws1, r, *item); r += 1
r += 2

# -- 3. 企业业务系统 --
r = write_sec_header(ws1, r, "三、企业业务系统", ACCENT_BLUE)
for item in [
    ("ERP  企业资源计划",
     "企业核心管理系统，打通财务、采购、库存、生产、销售、HR六大模块，实现数据出一门。药企版通常包含GMP/GSP合规模块。",
     "例：某生物药企上线ERP后，月结时间从7天缩短至2天，库存周转率提升35%",
     "选型重点：行业适配性（GMP/GLP）、实施服务商能力、二次开发灵活度",
     LIGHT_GREEN),
    ("CRM  客户关系管理",
     "管理从线索到商机到签约到服务全生命周期，提升销售转化率、客户复购率。药企常用于对接药企BD部门和临床合作方。",
     "例：某IVD企业CRM管理200加经销商，销售负责人随时查看商机进度，丢单率下降20%",
     "区分SFA（销售自动化）和营销自动化，避免买了大而无用的产品",
     LIGHT_GREEN),
    ("LIMS  实验室信息管理系统",
     "专为药企和检测机构设计，管理样品流转、检测方法、实验记录、报告生成，支持审计追踪（ALCOA加原则）。",
     "例：某CRO公司LIMS管理3000加样品每天，电子记录替代纸质记录迎检时间缩短60%",
     "无纸化办公趋势，NMPA/FDA审查必查；选有药品行业案例的服务商",
     LIGHT_GREEN),
    ("ELM  实验室执行系统",
     "管理研发实验全流程：实验设计（DOE）、实验记录、仪器连接、数据自动采集、配方管理。",
     "例：某疫苗研发团队用ELM管理小鼠实验，实验记录从4小时每份缩短至40分钟每份",
     "ELM与LIMS互补，LIMS管合规记录，ELM管实验过程",
     LIGHT_GREEN),
    ("QMS  质量管理系统",
     "管理CAPA、偏差、变更控制、供应商审计、验证管理等质量活动的全流程工具。",
     "例：某药企QMS上线后，偏差处理周期从45天缩短至18天，CAPA按时关闭率从62%提升至94%",
     "药企GMP核心系统，是审计高频考点",
     LIGHT_GREEN),
    ("MES  制造执行系统",
     "部署在生产车间，衔接ERP计划层与PLC/DCS控制层，实现生产过程透明化、药品追溯码关联。",
     "例：某生物制品企业MES实现车间实时产量监控，药品追溯码关联准确率99.99%",
     "2022年国家药监局要求药品追溯码全覆盖，MES是核心载体",
     LIGHT_GREEN),
]:
    write_prod_row(ws1, r, *item); r += 1
r += 2

# -- 4. 深信服 --
r = write_sec_header(ws1, r, "四、深信服 · 数据安全 & 网络安全（代理）", ACCENT_GREEN)
for item in [
    ("数据安全产品线",
     "包含：数据库审计（审计所有访问行为，识别拖库撞库风险）、数据脱敏（测试环境使用脱敏数据，防止泄密）、数据防泄漏DLP（敏感文件外发管控）。",
     "例：某药企研发数据曾险遭离职员工拷贝，DLP系统立即告警加阻断，保护了核心化合物数据",
     "等级保护/分保/商密认证的合规刚需",
     LIGHT_GREEN),
    ("网络安全产品线",
     "包含：下一代防火墙NGAF（应用层防护加IPS联动）、行为审计（内网全流量回溯）、VPN零信任（远程办公安全接入）、安全感知平台SIP（全局威胁可视化）。",
     "例：某药企被勒索病毒攻击，SIP平台2分钟内告警并自动隔离中毒主机，仅1台服务器受损",
     "一站式打包，兼容性好，运维成本低；我们提供从咨询到交付到驻场的全流程服务",
     LIGHT_GREEN),
]:
    write_prod_row(ws1, r, *item); r += 1
r += 2

# -- 5. 我们能做什么 --
r = write_sec_header(ws1, r, "五、我们能为你做什么", DARK_GRAY)
for item in [
    ("机房规划与咨询",
     "在机房建设初期介入，帮客户梳理IT现状与3-5年规划，出具机房建设规划建议书，避免建完就落后的结构性问题。",
     "免费现场勘查加初步方案 / 3-5年容量预测 / 预算估算",
     "客户收益：减少重复投资，避免扩建时业务中断",
     LIGHT_GRAY),
    ("机房装修设计与施工",
     "提供符合GB 50174标准的机房装修：天花/地板/隔断/配电/综合布线/消防/监控，一站式交钥匙工程。",
     "装饰装修 / 电气工程 / 弱电智能化 / 消防工程",
     "客户收益：统一管理责任，工期有保障，验收无忧",
     LIGHT_GRAY),
    ("设备选型与供应",
     "基于客户实际业务需求，提供国内外一线品牌服务器、存储、网络设备的选型参考与报价，提供设备采购服务。",
     "品牌：华为/戴尔/浪潮/新华三/深信服等 / 原厂授权 / 批量采购优惠",
     "客户收益：避开品牌雷区，减少试错成本",
     LIGHT_GRAY),
    ("实施部署与系统集成",
     "服务器系统安装、RAID配置、虚拟化平台部署（VMware/KVM）、业务系统安装调试、割接上线方案制定与执行。",
     "系统集成 / 数据迁移 / 7x24小时割接保障 / 应急回滚预案",
     "客户收益：专业实施，缩短上线周期，业务受损最小化",
     LIGHT_GRAY),
    ("运维服务（年度维保/驻场/SLA）",
     "提供多级运维服务：远程监控（动环监控加服务器状态）、故障响应（4小时到场/8小时解决）、定期巡检、备件先行更换。",
     "7x24小时监控值班 / 4小时到场（核心系统） / 季度巡检报告",
     "客户收益：IT团队减负，专注核心业务，系统可用性大于等于99.9%",
     LIGHT_GRAY),
]:
    write_prod_row(ws1, r, *item); r += 1

merge_write(ws1, r+2, 1, 5,
    "如有需求，欢迎联系我们！上海基数运维 - 您可信赖的IT基础设施合作伙伴",
    DARK_BLUE, fnt(True,11,WHITE), cal(), 32)

# ===================== Sheet 2 =====================
ws2 = wb.create_sheet("客户需求自检")
for i, w in enumerate([28,16,40,30], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

merge_write(ws2, 1, 1, 4, "客户需求自检表（销售填写）", DARK_BLUE, fnt(True,16,WHITE), cal(), 40)
merge_write(ws2, 2, 1, 4,
    "适用场景：新建实验室 / 实验室改造 / 机房改造 / 系统升级",
    MID_BLUE, fnt(False,10,WHITE,italic=True), cal(), 22)

r2 = 4
header_items = ["调查项目", "选项/填写", "说明", "客户填写"]
for i, h in enumerate(header_items, 1):
    c = ws2.cell(row=r2, column=i, value=h)
    c.fill=fill(DARK_BLUE); c.font=fnt(True,11,WHITE)
    c.alignment=cal(); c.border=bdr()
ws2.row_dimensions[r2].height = 25
r2 += 1

sections = [
    ("客户基本信息", ACCENT_BLUE, [
        ("客户名称", "文本填写", "公司全称", ""),
        ("所属行业", "请选择", "生物制药/医疗器械/CRO/CDMO/其他", ""),
        ("目前实验室/机房规模", "请选择", "小于100平/100-300平/300-500平/大于500平", ""),
    ]),
    ("机房基础设施现状", ACCENT_ORG, [
        ("现有服务器数量", "请选择", "小于5台/5-20台/20-50台/大于50台/暂无", ""),
        ("是否有专用机房", "请选择", "是（已投用___年）/否/计划建设中", ""),
        ("机柜数量", "请选择", "小于5个/5-20个/大于20个/暂无", ""),
        ("UPS配置情况", "请选择", "有（功率___kVA，续航___分钟）/无/不确定", ""),
        ("精密空调配置", "请选择", "有/无/普通空调凑合/不确定", ""),
    ]),
    ("IT系统现状", ACCENT_GREEN, [
        ("现有核心业务系统", "多选", "ERP/CRM/LIMS/ELM/QMS/MES/其他/暂无", ""),
        ("服务器存储品牌", "请填写", "戴尔/HPE/华为/浪潮/新华三/其他/暂无", ""),
        ("数据备份方式", "请选择", "磁带库/本地硬盘/云备份/无备份/不确定", ""),
        ("网络安全设备", "请选择", "防火墙/VPN/入侵检测/日志审计/无/不确定", ""),
    ]),
    ("未来3年规划（关键）", ACCENT_PURP, [
        ("预计新增服务器", "请选择", "小于10台/10-30台/30-50台/大于50台/未定", ""),
        ("计划上线新系统", "多选", "ERP/LIMS/MES/AI大数据/其他/未定", ""),
        ("算力需求", "请选择", "普通办公/中等算力（AI辅助）/高算力（AI训练）/暂无概念", ""),
        ("预算范围", "请选择", "50万以内/50-200万/200-500万/500万以上/暂无预算", ""),
    ]),
]

for sec_title, bg, items in sections:
    merge_write(ws2, r2, 1, 4, "  " + sec_title, bg, fnt(True,11,WHITE), lal(), 24)
    r2 += 1
    for row_data in items:
        for col, val in enumerate(row_data, 1):
            c = ws2.cell(row=r2, column=col, value=val)
            c.fill = fill(LIGHT_GRAY if r2%2==0 else WHITE)
            c.font = fnt(False, 10, DARK_GRAY)
            c.alignment = lal(); c.border = bdr()
        ws2.row_dimensions[r2].height = 28
        r2 += 1
    r2 += 1

merge_write(ws2, r2, 1, 4, "  销售备注（不限格式，自由填写）", DARK_BLUE, fnt(True,11,WHITE), lal(), 24)
r2 += 1
ws2.merge_cells(start_row=r2, start_column=1, end_row=r2+5, end_column=4)
note = ws2.cell(row=r2, column=1)
note.fill = fill("FFF9E6"); note.border = bdr("medium")
for ri in range(r2, r2+6):
    ws2.row_dimensions[ri].height = 20

# ===================== Sheet 3 =====================
ws3 = wb.create_sheet("为什么选我们")
for i, w in enumerate([6,30,50], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

merge_write(ws3, 1, 1, 3, "为什么选择上海基数运维", DARK_BLUE, fnt(True,16,WHITE), cal(), 40)

advantages = [
    (ACCENT_BLUE, "深耕IT基础设施20年",
     "核心团队来自华为、戴尔、联想，平均从业经验15年加，服务过100加生物制药客户，行业Know-How积累深厚"),
    (ACCENT_GREEN, "一站式产品供应",
     "服务器/存储/网络/安全/系统软件，原厂授权，品牌齐全，一站式采购，避免对接多个供应商的沟通成本"),
    (ACCENT_ORG, "自有施工资质与团队",
     "有机电工程施工总承包资质、装修资质、安全生产许可证，自有工程师加施工队伍，非层层外包"),
    (ACCENT_PURP, "专注生物制药行业",
     "服务过药明康德、君实生物、华大基因等知名药企，深刻理解GMP/GSP/LIMS合规要求"),
    (ACCENT_BLUE, "快速响应，本地化服务",
     "上海本地2小时到场，长三角4小时到场，7x24值班，全年无休，核心系统4小时到场保障"),
    (ACCENT_GREEN, "兄弟公司协同",
     "上海蓝西实验室设备（实验室整体建设）加上海基数运维（IT基础设施），资源互通，合作更紧密"),
]

r3 = 3
for bg, title, desc in advantages:
    merge_write(ws3, r3, 1, 3, "  " + title, bg, fnt(True,12,WHITE), lal(), 28)
    r3 += 1
    merge_write(ws3, r3, 1, 3, "  " + desc, LIGHT_GRAY if r3%2==0 else WHITE,
                fnt(False,10,DARK_GRAY), lal(), 30)
    r3 += 2

out_path = r"C:\Users\Administrator\Desktop\上海基数运维_合作伙伴产品手册.xlsx"
wb.save(out_path)
print("OK:", out_path)
