# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

DARK_BLUE   = "1E3A5F"
MID_BLUE    = "2C5F8A"
ACCENT_BLUE = "4A90D9"
ACCENT_GREEN= "27AE60"
ACCENT_ORG  = "E67E22"
LIGHT_BLUE  = "D6E4F0"
LIGHT_GREEN = "D5F5E3"
LIGHT_ORG   = "FDEBD0"
LIGHT_GRAY  = "F2F3F4"
WHITE       = "FFFFFF"
DARK_GRAY   = "2C3E50"
YELLOW_BG   = "FFF9E6"

def fill(c): return PatternFill("solid", fgColor=c)
def fnt(bold=False, size=10, color="000000", italic=False):
    return Font(name="微软雅黑", bold=bold, size=size, color=color, italic=italic)
def cal(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lal(): return Alignment(horizontal="left", vertical="center", wrap_text=True)
def bdr(style="thin"):
    s = Side(style=style, color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)
def bdr_med():
    s = Side(style="medium", color="2C3E50")
    return Border(left=s, right=s, top=s, bottom=s)

def mwrite(ws, row, c1, c2, value, bg, font_obj, aln_obj, h=30):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1, value=value)
    cell.fill=bg; cell.font=font_obj; cell.alignment=aln_obj; cell.border=bdr()
    ws.row_dimensions[row].height = h

def write_block(ws, start_row, icon_title, title, bg_header,
                 prob, prob_example, solve, solve_example, bg_row):
    """每块内容"""
    # 大标题行
    mwrite(ws, start_row, 1, 4, icon_title, bg_header, fnt(True,14,WHITE), lal(), 32)
    ws.row_dimensions[start_row].height = 32
    # 副标题（我们做什么）
    mwrite(ws, start_row+1, 1, 4, "  " + title, bg_header, fnt(False,10,WHITE,italic=True), lal(), 20)
    ws.row_dimensions[start_row+1].height = 20
    # 列标题
    for col, txt in [(1,"客户遇到的问题"), (2,"举个例子"), (3,"我们能帮他解决"), (4,"我们怎么帮")]:
        c = ws.cell(row=start_row+2, column=col, value=txt)
        c.fill=fill(DARK_GRAY); c.font=fnt(True,10,WHITE)
        c.alignment=cal(); c.border=bdr()
    ws.row_dimensions[start_row+2].height = 22
    # 内容行
    for row_i, (p, pe, s, se) in enumerate([(prob,prob_example,solve,solve_example)], start=start_row+3):
        bg = bg_row if row_i%2==0 else WHITE
        for col, val in [(1,p), (2,pe), (3,s), (4,se)]:
            c = ws.cell(row=row_i, column=col, value=val)
            c.fill=fill(bg); c.font=fnt(False,10,DARK_GRAY)
            c.alignment=lal(); c.border=bdr()
        ws.row_dimensions[row_i].height = 50
    return start_row + 5

# ===================== 主工作表 =====================
ws = wb.active
ws.title = "一句话介绍我能做什么"

# 列宽
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 26
ws.column_dimensions["C"].width = 26
ws.column_dimensions["D"].width = 26

# 主标题
ws.merge_cells("A1:D1")
t = ws["A1"]
t.value = "上海基数运维 · 一句话告诉客户我们能做什么"
t.fill = fill(DARK_BLUE)
t.font = fnt(True,18,WHITE)
t.alignment = cal()
ws.row_dimensions[1].height = 50

ws.merge_cells("A2:D2")
s = ws["A2"]
s.value = "蓝西销售专用版 · 客户问起来，3分钟讲清楚"
s.fill = fill(MID_BLUE)
s.font = fnt(False,10,WHITE,italic=True)
s.alignment = cal()
ws.row_dimensions[2].height = 22

# 引导说明
ws.merge_cells("A3:D3")
g = ws["A3"]
g.value = (
    "使用说明：左侧是客户常问的问题，中间是我们的解决方案。客户问什么，你指哪行念哪行，3分钟让客户知道我们能帮他解决。\n"
    "合作模式：上海基数运维（IT基础设施）+ 上海蓝西实验室设备（实验室整体建设），兄弟联手，客户的IT问题我们全包"
)
g.fill = fill(YELLOW_BG)
g.font = fnt(False,9,DARK_GRAY)
g.alignment = lal()
g.border = bdr_med()
ws.row_dimensions[3].height = 42

r = 5

# ══════════════════════════════════════════════
r = write_block(ws, r, "🏗  场景一：客户刚做完实验室装修，接下来IT怎么办？",
                "机房建设 / 服务器采购 / 网络安全 · 交钥匙工程",
                DARK_BLUE,
                "客户刚盖好实验室，装修完了，但IT这块完全没规划，不知道要买多少服务器、怎么配网络、以后怎么扩展",
                "某生物药企：实验室建好了，才发现机房没规划，增量改造花了2倍的钱和半年时间",
                "我们帮他做IT整体规划，出方案、供设备、搞实施，一站式搞定",
                "上门勘查 → 出规划方案 → 供应设备 → 部署实施 → 后续运维，客户只对接我们一个供应商",
                LIGHT_BLUE)

# ══════════════════════════════════════════════
r = write_block(ws, r, "💻  场景二：客户要上ERP/LIMS等业务系统，但没有服务器",
                "服务器采购 / 系统部署 · 配套服务",
                ACCENT_BLUE,
                "客户想买ERP/LIMS等系统，软件商说需要服务器，但客户不知道买什么配置、买几台",
                "某IVD公司：买了LIMS软件，IT说需要服务器，不知道怎么选，配了3个月才配好，耽误上线",
                "我们帮他选型、采购、安装、调试，客户坐等可用",
                "告诉客户用途和预算 → 我们出配置清单 → 供货 → 上门安装调试 → 交付使用",
                LIGHT_BLUE)

# ══════════════════════════════════════════════
r = write_block(ws, r, "🤖  场景三：客户想做AI，但完全不懂算力",
                "AI算力服务器 · 药物研发加速",
                ACCENT_ORG,
                "客户听说AI很火，想用AI做药物研发、化合物筛选等，但不知道需要什么GPU服务器，不知道多少钱",
                "某 biotech 公司：想用AI筛选化合物，买了普通服务器跑不动，才知道要买GPU服务器，重复采购花了冤枉钱",
                "我们根据他的AI场景推荐GPU服务器型号和数量，让他一次买对",
                "沟通AI应用场景 → 推荐合适的GPU服务器 → 供货+部署 → 效果验证",
                LIGHT_ORG)

# ══════════════════════════════════════════════
r = write_block(ws, r, "🔒  场景四：客户担心数据安全，怕被攻击/泄密",
                "网络安全 / 数据安全 · 深信服代理",
                ACCENT_GREEN,
                "客户担心服务器被黑客攻击、数据被偷、员工泄密，但不知道怎么防护",
                "某药企：研发数据差点被离职员工拷走幸好发现及时；一次勒索病毒让整个IT系统瘫痪2周",
                "我们代理深信服全套产品，防火墙+数据防泄漏+日志审计，一套方案全部搞定",
                "评估安全现状 → 出方案 → 供货+部署 → 7x24安全监控运维",
                LIGHT_GREEN)

# ══════════════════════════════════════════════
r = write_block(ws, r, "📈  场景五：客户业务发展快，服务器不够用了",
                "服务器扩容 · 平滑升级",
                "2986C1",
                "客户业务增长快，服务器跑不动了，存储不够了，想扩容但不知道怎么扩、会不会影响现有业务",
                "某药企：业务快速发展，服务器跑不动，扩容时业务中断了一整天，丢了不少订单",
                "我们帮他评估现状，规划扩容方案，平滑升级不断服",
                "现场评估 → 制定扩容方案 → 利旧现有设备 → 平滑升级 → 业务不中断",
                LIGHT_BLUE)

# ══════════════════════════════════════════════
r = write_block(ws, r, "🛠  场景六：客户IT没人维护，出了问题没人管",
                "IT运维服务 · 全包托管",
                DARK_GRAY,
                "客户IT团队只有1-2个人，忙不过来；或者根本没有IT，服务器出了问题到处找人",
                "某初创药企：IT就一个兼职，服务器崩溃凌晨2点找不到人，业务停了2天",
                "我们提供全年IT托管服务，远程监控+故障到场+定期巡检，客户专注业务就行",
                "签约年度维保 → 7x24监控 → 故障4小时到场 → 季度巡检报告",
                LIGHT_GRAY)

# 结尾
ws.merge_cells(f"A{r}:D{r}")
c = ws.cell(row=r, column=1,
    value="📞 有任何客户IT需求，直接联系我们！上海基数运维 · 上海蓝西实验室设备 · 兄弟联手，为客户提供一站式服务")
c.fill=fill(DARK_BLUE); c.font=fnt(True,10,WHITE)
c.alignment=cal(); c.border=bdr_med()
ws.row_dimensions[r].height = 36

# ===================== Sheet2: 快速参考卡 =====================
ws2 = wb.create_sheet("快速参考")

ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 40
ws2.column_dimensions["C"].width = 40

ws2.merge_cells("A1:C1")
t2 = ws2["A1"]
t2.value = "快速参考卡 · 3分钟背完"
t2.fill = fill(DARK_BLUE)
t2.font = fnt(True,16,WHITE)
t2.alignment = cal()
ws2.row_dimensions[1].height = 42

cards = [
    ("我能做什么（6句话）", [
        "① 帮客户做机房IT整体规划，建完不后悔",
        "② 卖服务器、存储、网络设备，一线品牌都有",
        "③ 帮客户部署ERP/LIMS等系统，配好就能用",
        "④ 卖AI算力服务器，帮 biotech 用上AI研发",
        "⑤ 深信服网络安全，数据防住、攻击挡住",
        "⑥ IT全年托管运维，客户只管专注业务",
    ]),
    ("客户最常问的6个问题", [
        "Q1：机房要怎么建？ → 找我们，出方案+施工+设备一条龙",
        "Q2：服务器买几台？ → 告诉我们你上什么系统，我们帮你算",
        "Q3：想做AI要多少预算？→ 看场景，小则10万，大则上百万，我们给你选型",
        "Q4：数据安全吗？→ 深信服全线产品，防火墙+数据防泄漏，合规等保",
        "Q5：以后要扩展怎么办？→ 规划阶段就帮你考虑扩容，避免重复建设",
        "Q6：有售后吗？→ 有，7x24值班，核心系统4小时到场",
    ]),
    ("我们的优势（3句话）", [
        "1. 深耕IT 20年，服务过100+生物制药客户，行业懂行",
        "2. 服务器/存储/网络/安全/系统，一站式供齐，不用对多个供应商",
        "3. 兄弟公司上海蓝西（实验室建设）+ 我们（IT基础设施），客户IT问题全覆盖",
    ]),
    ("什么时候找我们（4个时机）", [
        "✓ 客户刚做完实验室/厂房装修 → 趁机房还空，现在介入最合适",
        "✓ 客户要上ERP/LIMS等系统 → 服务器选型采购交给我们",
        "✓ 客户担心数据安全/怕攻击 → 深信服代理，给客户打包方案",
        "✓ 客户IT团队人手不够 → 年度托管，客户减负专注业务",
    ]),
]

r2 = 3
for sec_title, items in cards:
    ws2.merge_cells(f"A{r2}:C{r2}")
    sc = ws2.cell(row=r2, column=1, value=sec_title)
    sc.fill=fill(ACCENT_BLUE); sc.font=fnt(True,11,WHITE)
    sc.alignment=lal(); sc.border=bdr()
    ws2.row_dimensions[r2].height = 26
    r2 += 1
    for item in items:
        for col, val in [(1,""), (2,item), (3,"")]:
            c = ws2.cell(row=r2, column=col, value=val)
            c.fill=fill(LIGHT_GRAY if r2%2==0 else WHITE)
            c.font=fnt(False,10,DARK_GRAY)
            c.alignment=lal(); c.border=bdr()
        ws2.row_dimensions[r2].height = 24
        r2 += 1
    r2 += 1

out = r"C:\Users\Administrator\Desktop\上海基数运维_销售简单版.xlsx"
wb.save(out)
print("OK:", out)
